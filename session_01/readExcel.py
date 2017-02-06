# Javier Vazquez
# Python 3.6

import openpyxl, os, re, pprint, sys, logging

os.remove("log.txt")
logging.basicConfig(filename="log.txt", level = logging.DEBUG, format =' %(asctime)s - %(levelname)s - %(message)s')

class excelOperations:

    def __init__(self):
        self.__wb = 0
        self.__ws = 0
        self.__excelDict = {}

    @property
    def wb(self):
        return self.__wb

    @property
    def ws(self):
        return self.__ws

    @wb.setter
    def wb(self, wb):
        self.__wb = wb

    @ws.setter
    def ws(self, ws):
        self.__ws = ws

    def displayExcelFiles(self):
        '''Display all the excel files'''
        listXlsx = []
        extRegex = re.compile(r".+xlsx$")
        listFiles = os.listdir()
        for file in listFiles:
            ext = extRegex.search(file)
            if ext is not None:
                listXlsx.append(ext.group())
        return listXlsx

    def readExcelFile(self, file):
        '''Read an excel file'''
        if os.path.isfile(file):
            self.__wb = openpyxl.load_workbook(file)
            self.__ws = self.__wb.active
            return self.__wb, self.__ws
        else:
            return False, False

    def displayContent(self):
        '''Display the content of a excel file'''
        for i in range(1, self.__ws.max_row + 1):
            rowToDisplay = ""
            for j in range(1, self.__ws.max_column + 1):
                rowToDisplay += str(self.__ws.cell(row=i, column=j).value) + " "
            print(str(i)+": "+rowToDisplay)
            logging.debug("Row: " + rowToDisplay)

    def createDictionary(self):
        '''Create dictionary from an excel file'''
        for i in range(2, self.__ws.max_row + 1):
            produce = self.__ws.cell(row=i, column=1).value
            if produce not in self.__excelDict.keys():
                self.__excelDict[produce] = self.__ws.cell(row=i, column=2).value
                logging.debug("Item added to dict- "+str(produce)+": "+str(self.__excelDict[produce]))

    def saveDictionary(self, mode):
        line = ""
        textFile = open("dictionary.txt", mode)
        # sortedDictList = sorted(self.__excelDict.items(), reverse = False, key=lambda x: x[0])
        for key, value in self.__excelDict.items():
            line += str(key) + ": " + str(value) + "\n"
        textFile.write(line)
        textFile.close()

    def loadDictionary(self):
        self.__excelDict = {}
        produceRegex = re.compile(r"^(\w.+):")
        priceRegex = re.compile(r":(.+)$")
        textFile = open("dictionary.txt", "r")
        textLines = textFile.readlines()
        for textLine in textLines:
            matchingProduce = produceRegex.search(textLine)
            matchingPrice = priceRegex.search(textLine)
            if matchingProduce.group(1) not in self.__excelDict.keys():
                self.__excelDict[matchingProduce.group(1)] = float(matchingPrice.group(1))
        textFile.close()

    def printDictionary(self):
        sortedDictList = sorted(self.__excelDict.items(), reverse=False, key=lambda x: x[0])
        for i in range(0, len(sortedDictList), 2):
            print("{0:<20}{1:>5}{2:<5}{3:<20}{4:>5}"
                  .format(sortedDictList[i][0], sortedDictList[i][1], " ",
                          sortedDictList[i+1][0], sortedDictList[i+1][1]))

    def updateKey(self, key, value):
        if key in self.__excelDict.keys():
            self.__excelDict[key] = value
            self.saveDictionary("w")
            return True
        else:
            return False

    def updateExcel(self, key, value):
        for i in range(2, self.__ws.max_row + 1):
        # for i in range(2, 1000 + 1):
            if self.__ws.cell(row=i, column=1).value == key:
                self.__ws.cell(row=i, column=2).value = float(value)
                self.__wb.save("produceSales.xlsx")

    def sortByLabel(self):
        '''Sort every value by its label'''
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "countBySex"
        f = m = 2
        headers = openpyxl.styles.Font(bold=True)
        ws.cell(row=1, column=1).font = headers
        ws.cell(row=1, column=2).font = headers
        ws.cell(row=1, column=1).value = "Female"
        ws.cell(row=1, column=2).value = "Male"
        for i in range(2, self.__ws.max_row + 1):
            if self.__ws.cell(row=i, column=3).value == "Female":
                ws.cell(row=f, column=1).value = self.__ws.cell(row=i, column=1).value * 12 + self.__ws.cell(row=i, column=2).value
                f += 1
            if self.__ws.cell(row=i, column=3).value == "Male":
                ws.cell(row=m, column=2).value = self.__ws.cell(row=i, column=1).value * 12 + self.__ws.cell(row=i, column=2).value
                m += 1
        print("Females: ", f-2)
        print("Males: ", m-2)
        wb.save("assigment_01.xlsx")

    def excelColsToList(self, n):
        '''List per row, n is the sample size if n is max is all the population'''
        female = []
        male = []
        m=0
        if n.isalpha():
            if n == "max":
                m = self.__ws.max_row + 1
            else:
                return False, False
        elif n.isdigit():
            m = int(n) + 2
        for j in range(2, m):
            if self.__ws.cell(row=j, column=1).value is not None:
                female.append(self.__ws.cell(row=j, column=1).value)
        for j in range(2, m):
            if self.__ws.cell(row=j, column=2).value is not None:
                male.append(self.__ws.cell(row=j, column=2).value)
        return female, male

if __name__ == '__main__':
    print("Direct access to "+ os.path.basename(__file__))
else:
    print(os.path.basename(__file__)+" class instance")
